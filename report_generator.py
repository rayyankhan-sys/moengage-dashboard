"""
Excel Report Generator for MoEngage Dashboard
Creates Excel reports with raw data and computed metrics using openpyxl
"""
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import MoEngageDatabase
from config import COUNTRIES, PN_SENT_TO_IMPRESSION_RATIO

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates Excel reports with metrics"""

    def __init__(self):
        """Initialize report generator"""
        self.db = MoEngageDatabase()

    # ========================================================================
    # STYLE HELPERS
    # ========================================================================

    def _get_header_style(self) -> Font:
        """Get bold font for headers"""
        return Font(bold=True, size=11)

    def _get_section_header_fill(self) -> PatternFill:
        """Get gray background for section headers"""
        return PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    def _format_percentage(self, value: float) -> str:
        """Format value as percentage with 2 decimals"""
        if value is None or value == 0:
            return "0.00%"
        return f"{value:.2f}%"

    def _format_number(self, value: int) -> str:
        """Format number with thousands separator"""
        if value is None:
            return "0"
        return f"{value:,}"

    # ========================================================================
    # METRIC CALCULATION FUNCTIONS
    # ========================================================================

    def calculate_comms_reachable_pct_total(
        self,
        reachable: int,
        total_users: int,
    ) -> float:
        """Comms Reachable (% Total) = reachable / total_users * 100"""
        if total_users == 0:
            return 0.0
        return (reachable / total_users) * 100

    def calculate_comms_reached_pct_total(
        self,
        received: int,
        total_users: int,
    ) -> float:
        """Comms Reached (% Total) = received / total_users * 100"""
        if total_users == 0:
            return 0.0
        return (received / total_users) * 100

    def calculate_comms_reachable_pct_active(
        self,
        reachable: int,
        active_users: int,
    ) -> float:
        """Comms Reachable (% Active) = reachable / active_users * 100"""
        if active_users == 0:
            return 0.0
        return (reachable / active_users) * 100

    def calculate_comms_reached_pct_active(
        self,
        received: int,
        active_users: int,
    ) -> float:
        """Comms Reached (% Active) = received / active_users * 100"""
        if active_users == 0:
            return 0.0
        return (received / active_users) * 100

    def calculate_unsubscribe_rate(
        self,
        unsubscribed: int,
        total_users: int,
    ) -> float:
        """Unsubscribe Rate = unsubscribed / total_users * 100"""
        if total_users == 0:
            return 0.0
        return (unsubscribed / total_users) * 100

    def calculate_est_push_sent(
        self,
        impressions: int,
        pn_ratio: float,
    ) -> float:
        """Est. Promo Push Sent = impressions * pn_ratio"""
        return impressions * pn_ratio

    def calculate_avg_comms_per_user(
        self,
        comms_sent: int,
        users: int,
    ) -> float:
        """Avg Comms per User = comms_sent / users"""
        if users == 0:
            return 0.0
        return comms_sent / users

    def calculate_ctr_impression_basis(
        self,
        clicks: int,
        impressions: int,
    ) -> float:
        """CTR (Impression basis) = clicks / impressions * 100"""
        if impressions == 0:
            return 0.0
        return (clicks / impressions) * 100

    def calculate_ctr_sent_basis(
        self,
        clicks: int,
        sent: int,
    ) -> float:
        """CTR (Sent basis) = clicks / sent * 100"""
        if sent == 0:
            return 0.0
        return (clicks / sent) * 100

    def calculate_email_open_rate(
        self,
        opens: int,
        sent: int,
    ) -> float:
        """Email Open Rate = opens / sent * 100"""
        if sent == 0:
            return 0.0
        return (opens / sent) * 100

    def calculate_email_ctr(
        self,
        clicks: int,
        sent: int,
    ) -> float:
        """Email CTR = clicks / sent * 100"""
        if sent == 0:
            return 0.0
        return (clicks / sent) * 100

    def calculate_comms_per_reachable_user(
        self,
        comms_sent: int,
        reachable: int,
    ) -> float:
        """Comms per Reachable User = comms_sent / reachable"""
        if reachable == 0:
            return 0.0
        return comms_sent / reachable

    # ========================================================================
    # REPORT GENERATION
    # ========================================================================

    def generate_report(
        self,
        period_start: str,
        period_end: str,
        comparison_start: Optional[str] = None,
        comparison_end: Optional[str] = None,
    ) -> BytesIO:
        """
        Generate Excel report with metrics
        Returns: BytesIO object containing the workbook
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Fetch data
        segment_metrics = self.db.get_all_segment_metrics(period_start, period_end)
        campaign_metrics = self.db.get_all_campaign_metrics(period_start, period_end)

        comparison_segments = None
        comparison_campaigns = None
        if comparison_start and comparison_end:
            comparison_segments = self.db.get_all_segment_metrics(
                comparison_start, comparison_end
            )
            comparison_campaigns = self.db.get_all_campaign_metrics(
                comparison_start, comparison_end
            )

        # Create sheets
        self._add_summary_sheet(
            wb,
            "Summary — UK",
            "GB",
            segment_metrics,
            campaign_metrics,
            comparison_segments,
            comparison_campaigns,
        )

        self._add_summary_sheet(
            wb,
            "Summary — UAE",
            "AE",
            segment_metrics,
            campaign_metrics,
            comparison_segments,
            comparison_campaigns,
        )

        self._add_segments_sheet(wb, segment_metrics, comparison_segments)

        self._add_campaigns_sheet(wb, campaign_metrics, comparison_campaigns)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _add_summary_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        country: str,
        segment_metrics: List[Dict],
        campaign_metrics: List[Dict],
        comparison_segments: Optional[List[Dict]],
        comparison_campaigns: Optional[List[Dict]],
    ) -> None:
        """Add summary metrics sheet for a specific country"""
        ws = wb.create_sheet(sheet_name)

        # Create lookup dicts
        segment_dict = {
            (s["segment_type"], s["country"]): s.get("user_count", 0)
            for s in segment_metrics
        }
        campaign_dict = {
            (c["country"], c["channel"], c["campaign_type"]): c
            for c in campaign_metrics
        }

        if comparison_segments:
            comp_segment_dict = {
                (s["segment_type"], s["country"]): s.get("user_count", 0)
                for s in comparison_segments
            }
        else:
            comp_segment_dict = {}

        # Get PN ratio for this country
        pn_ratio = PN_SENT_TO_IMPRESSION_RATIO.get(country, 1.0)

        row = 1

        # ====================================================================
        # RAW DATA — User Base
        # ====================================================================
        self._add_section_header(ws, row, "RAW DATA — User Base")
        row += 1

        total_users = segment_dict.get(("TOTAL_USERS", country), 0)
        row = self._add_metric_row(ws, row, "Total Users", total_users, country, segment_dict, comp_segment_dict, "TOTAL_USERS")

        active_users = segment_dict.get(("ACTIVE_USERS_60D", country), 0)
        row = self._add_metric_row(ws, row, "Active Users (60d)", active_users, country, segment_dict, comp_segment_dict, "ACTIVE_USERS_60D")

        transacted_users = segment_dict.get(("TRANSACTED_USERS_PERIOD", country), 0)
        row = self._add_metric_row(ws, row, "Transacted Users", transacted_users, country, segment_dict, comp_segment_dict, "TRANSACTED_USERS_PERIOD")

        row += 1

        # ====================================================================
        # RAW DATA — Segment Counts
        # ====================================================================
        self._add_section_header(ws, row, "RAW DATA — Segment Counts")
        row += 1

        # FIX BUG 2: Use correct segment type keys matching data_puller.py
        received_push = segment_dict.get(("RECEIVED_PUSH_PERIOD", country), 0)
        row = self._add_metric_row(ws, row, "Unique Users Received Push", received_push, country, segment_dict, comp_segment_dict, "RECEIVED_PUSH_PERIOD")

        received_email = segment_dict.get(("RECEIVED_EMAIL_PERIOD", country), 0)
        row = self._add_metric_row(ws, row, "Unique Users Received Email", received_email, country, segment_dict, comp_segment_dict, "RECEIVED_EMAIL_PERIOD")

        active_received_email = segment_dict.get(("ACTIVE_EMAIL_PERIOD", country), 0)
        row = self._add_metric_row(ws, row, "Active Users Received Email", active_received_email, country, segment_dict, comp_segment_dict, "ACTIVE_EMAIL_PERIOD")

        unsub_push = segment_dict.get(("UNSUBSCRIBED_PUSH_PERIOD", country), 0)
        push_reachable = total_users - unsub_push
        row = self._add_computed_metric_row(ws, row, "Push Reachable (Raw)", push_reachable)

        unsub_email = segment_dict.get(("UNSUBSCRIBED_EMAIL_PERIOD", country), 0)
        email_reachable = total_users - unsub_email
        row = self._add_computed_metric_row(ws, row, "Email Reachable (Raw)", email_reachable)

        row = self._add_metric_row(ws, row, "Unsubscribed Push", unsub_push, country, segment_dict, comp_segment_dict, "UNSUBSCRIBED_PUSH_PERIOD")

        row = self._add_metric_row(ws, row, "Unsubscribed Email", unsub_email, country, segment_dict, comp_segment_dict, "UNSUBSCRIBED_EMAIL_PERIOD")

        row += 1

        # ====================================================================
        # RAW DATA — Promotional
        # ====================================================================
        self._add_section_header(ws, row, "RAW DATA — Promotional")
        row += 1

        promo_push_campaigns = [
            c for c in campaign_metrics
            if c["country"] == country and c["channel"] == "push" and c["campaign_type"] == "promotional"
        ]
        promo_email_campaigns = [
            c for c in campaign_metrics
            if c["country"] == country and c["channel"] == "email" and c["campaign_type"] == "promotional"
        ]

        # FIX BUG 3: Use "sent" field for push impressions (matching dashboard.py)
        # MoEngage push stats "sent" = impressions delivered to device
        promo_push_impressions = sum(c.get("sent", 0) for c in promo_push_campaigns)
        row = self._add_computed_metric_row(ws, row, "Promo Push Impressions", promo_push_impressions)

        promo_push_clicks = sum(c.get("click", 0) for c in promo_push_campaigns)
        row = self._add_computed_metric_row(ws, row, "Promo Push Clicks", promo_push_clicks)

        promo_email_sent = sum(c.get("sent", 0) for c in promo_email_campaigns)
        row = self._add_computed_metric_row(ws, row, "Promo Email Sent", promo_email_sent)

        promo_email_opens = sum(c.get("open", 0) for c in promo_email_campaigns)
        row = self._add_computed_metric_row(ws, row, "Promo Email Opens", promo_email_opens)

        promo_email_clicks = sum(c.get("click", 0) for c in promo_email_campaigns)
        row = self._add_computed_metric_row(ws, row, "Promo Email Clicks", promo_email_clicks)

        push_campaign_count = len(promo_push_campaigns)
        row = self._add_computed_metric_row(ws, row, "Push Campaigns count", push_campaign_count)

        email_campaign_count = len(promo_email_campaigns)
        row = self._add_computed_metric_row(ws, row, "Email Campaigns count", email_campaign_count)

        ws.cell(row, 1, "PN Sent:Impression Ratio")
        ws.cell(row, 2, pn_ratio)
        row += 1

        row += 1

        # ====================================================================
        # RAW DATA — Transactional
        # ====================================================================
        self._add_section_header(ws, row, "RAW DATA — Transactional")
        row += 1

        txn_push_campaigns = [
            c for c in campaign_metrics
            if c["country"] == country and c["channel"] == "push" and c["campaign_type"] == "transactional"
        ]
        txn_email_campaigns = [
            c for c in campaign_metrics
            if c["country"] == country and c["channel"] == "email" and c["campaign_type"] == "transactional"
        ]

        txn_push_sent = sum(c.get("sent", 0) for c in txn_push_campaigns)
        row = self._add_computed_metric_row(ws, row, "Transactional Push Sent", txn_push_sent)

        txn_email_sent = sum(c.get("sent", 0) for c in txn_email_campaigns)
        row = self._add_computed_metric_row(ws, row, "Transactional Email Sent", txn_email_sent)

        row += 1

        # ====================================================================
        # COMPUTED METRICS — Reachability & Reach
        # ====================================================================
        self._add_section_header(ws, row, "COMPUTED METRICS — Reachability & Reach")
        row += 1

        if total_users > 0:
            push_reachable_pct = self.calculate_comms_reachable_pct_total(push_reachable, total_users)
            ws.cell(row, 1, "Comms Reachable (% Total) — Push")
            ws.cell(row, 2, round(push_reachable_pct, 2))
            ws.cell(row, 2).number_format = '0.00"%"'
            row += 1

            email_reachable_pct = self.calculate_comms_reachable_pct_total(email_reachable, total_users)
            ws.cell(row, 1, "Comms Reachable (% Total) — Email")
            ws.cell(row, 2, round(email_reachable_pct, 2))
            ws.cell(row, 2).number_format = '0.00"%"'
            row += 1

            push_reached_pct = self.calculate_comms_reached_pct_total(received_push, total_users)
            ws.cell(row, 1, "Comms Reached (% Total) — Push")
            ws.cell(row, 2, round(push_reached_pct, 2))
            ws.cell(row, 2).number_format = '0.00"%"'
            row += 1

            email_reached_pct = self.calculate_comms_reached_pct_total(received_email, total_users)
            ws.cell(row, 1, "Comms Reached (% Total) — Email")
            ws.cell(row, 2, round(email_reached_pct, 2))
            ws.cell(row, 2).number_format = '0.00"%"'
            row += 1

        if active_users > 0:
            push_reachable_active_pct = self.calculate_comms_reachable_pct_active(push_reachable, active_users)
            ws.cell(row, 1, "Comms Reachable (% Active) — Push")
            ws.cell(row, 2, round(push_reachable_active_pct, 2))
            ws.cell(row, 2).number_format = '0.00"%"'
            row += 1

            email_reached_active_pct = self.calculate_comms_reached_pct_active(active_received_email, active_users)
            ws.cell(row, 1, "Comms Reached (% Active) — Email")
            ws.cell(row, 2, round(email_reached_active_pct, 2))
            ws.cell(row, 2).number_format = '0.00"%"'
            row += 1

        row += 1

        # ====================================================================
        # COMPUTED METRICS — Unsubscribes
        # ====================================================================
        self._add_section_header(ws, row, "COMPUTED METRICS — Unsubscribes")
        row += 1

        if total_users > 0:
            unsub_push_rate = self.calculate_unsubscribe_rate(unsub_push, total_users)
            ws.cell(row, 1, "Unsubscribe Rate — Push")
            ws.cell(row, 2, round(unsub_push_rate, 2))
            ws.cell(row, 2).number_format = '0.00"%"'
            row += 1

            unsub_email_rate = self.calculate_unsubscribe_rate(unsub_email, total_users)
            ws.cell(row, 1, "Unsubscribe Rate — Email")
            ws.cell(row, 2, round(unsub_email_rate, 2))
            ws.cell(row, 2).number_format = '0.00"%"'
            row += 1

        row += 1

        # ====================================================================
        # COMPUTED METRICS — Est. Promo Push Sent
        # ====================================================================
        self._add_section_header(ws, row, "COMPUTED METRICS — Est. Promo Push Sent")
        row += 1

        est_promo_push_sent = self.calculate_est_push_sent(promo_push_impressions, pn_ratio)
        ws.cell(row, 1, "Est. Promo Push Sent")
        ws.cell(row, 2, round(est_promo_push_sent, 2))
        row += 1

        row += 1

        # ====================================================================
        # COMPUTED METRICS — Avg Comms per User
        # ====================================================================
        self._add_section_header(ws, row, "COMPUTED METRICS — Avg Comms per User")
        row += 1

        if total_users > 0:
            avg_promo_push = self.calculate_avg_comms_per_user(est_promo_push_sent, total_users)
            ws.cell(row, 1, "Avg Promo Push per User")
            ws.cell(row, 2, round(avg_promo_push, 2))
            row += 1

            avg_promo_email = self.calculate_avg_comms_per_user(promo_email_sent, total_users)
            ws.cell(row, 1, "Avg Promo Email per User")
            ws.cell(row, 2, round(avg_promo_email, 2))
            row += 1

        if transacted_users > 0:
            avg_txn_push = self.calculate_avg_comms_per_user(txn_push_sent, transacted_users)
            ws.cell(row, 1, "Avg Txn Push per User")
            ws.cell(row, 2, round(avg_txn_push, 2))
            row += 1

            avg_txn_email = self.calculate_avg_comms_per_user(txn_email_sent, transacted_users)
            ws.cell(row, 1, "Avg Txn Email per User")
            ws.cell(row, 2, round(avg_txn_email, 2))
            row += 1

        row += 1

        # ====================================================================
        # COMPUTED METRICS — Performance
        # ====================================================================
        self._add_section_header(ws, row, "COMPUTED METRICS — Performance")
        row += 1

        promo_push_ctr_impression = self.calculate_ctr_impression_basis(promo_push_clicks, promo_push_impressions)
        ws.cell(row, 1, "Promo Push CTR (Impression basis)")
        ws.cell(row, 2, round(promo_push_ctr_impression, 2))
        ws.cell(row, 2).number_format = '0.00"%"'
        row += 1

        promo_push_ctr_sent = self.calculate_ctr_sent_basis(promo_push_clicks, est_promo_push_sent)
        ws.cell(row, 1, "Promo Push CTR (Est. Sent basis)")
        ws.cell(row, 2, round(promo_push_ctr_sent, 2))
        ws.cell(row, 2).number_format = '0.00"%"'
        row += 1

        email_open_rate = self.calculate_email_open_rate(promo_email_opens, promo_email_sent)
        ws.cell(row, 1, "Email Open Rate")
        ws.cell(row, 2, round(email_open_rate, 2))
        ws.cell(row, 2).number_format = '0.00"%"'
        row += 1

        promo_email_ctr = self.calculate_email_ctr(promo_email_clicks, promo_email_sent)
        ws.cell(row, 1, "Promo Email CTR")
        ws.cell(row, 2, round(promo_email_ctr, 2))
        ws.cell(row, 2).number_format = '0.00"%"'
        row += 1

        row += 1

        # ====================================================================
        # COMPUTED METRICS — PNs & Emails per Reachable User
        # ====================================================================
        self._add_section_header(ws, row, "COMPUTED METRICS — PNs & Emails per Reachable User")
        row += 1

        total_push_sent = est_promo_push_sent + txn_push_sent
        pns_per_reachable = self.calculate_comms_per_reachable_user(total_push_sent, push_reachable)
        ws.cell(row, 1, "PNs per Push Reachable User")
        ws.cell(row, 2, round(pns_per_reachable, 2))
        row += 1

        total_email_sent = promo_email_sent + txn_email_sent
        emails_per_reachable = self.calculate_comms_per_reachable_user(total_email_sent, email_reachable)
        ws.cell(row, 1, "Emails per Email Reachable User")
        ws.cell(row, 2, round(emails_per_reachable, 2))
        row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20

    def _add_section_header(self, ws, row: int, text: str) -> None:
        """Add a section header row"""
        ws.cell(row, 1, text)
        ws.cell(row, 1).font = self._get_header_style()
        ws.cell(row, 1).fill = self._get_section_header_fill()

    def _add_metric_row(
        self,
        ws,
        row: int,
        label: str,
        value: int,
        country: str,
        segment_dict: Dict,
        comp_segment_dict: Dict,
        segment_type: str,
    ) -> int:
        """Add a metric row with optional comparison columns"""
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.cell(row, 2).number_format = '#,##0'

        if comp_segment_dict:
            comp_value = comp_segment_dict.get((segment_type, country), 0)
            ws.cell(row, 3, comp_value)
            ws.cell(row, 3).number_format = '#,##0'

            if comp_value > 0:
                change_pct = ((value - comp_value) / comp_value) * 100
                ws.cell(row, 4, round(change_pct, 2))
                ws.cell(row, 4).number_format = '0.00"%"'

        return row + 1

    def _add_computed_metric_row(
        self,
        ws,
        row: int,
        label: str,
        value: float,
    ) -> int:
        """Add a computed metric row"""
        ws.cell(row, 1, label)
        if isinstance(value, float):
            ws.cell(row, 2, round(value, 2))
        else:
            ws.cell(row, 2, value)
        ws.cell(row, 2).number_format = '#,##0'

        return row + 1

    def _add_segments_sheet(
        self,
        wb: Workbook,
        segment_metrics: List[Dict],
        comparison_segments: Optional[List[Dict]],
    ) -> None:
        """Add segments data sheet"""
        ws = wb.create_sheet("Segments")

        # Headers
        headers = ["Segment Type", "Country", "User Count"]
        if comparison_segments:
            headers.extend(["Previous", "Change (%)"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = self._get_header_style()
            cell.fill = self._get_section_header_fill()

        # Data rows
        row = 2
        for metric in segment_metrics:
            ws.cell(row, 1, metric.get("segment_type"))
            ws.cell(row, 2, metric.get("country"))
            ws.cell(row, 3, metric.get("user_count", 0))
            ws.cell(row, 3).number_format = '#,##0'

            if comparison_segments:
                # Find comparison metric
                comparison = next(
                    (
                        s
                        for s in comparison_segments
                        if s["segment_type"] == metric["segment_type"]
                        and s["country"] == metric["country"]
                    ),
                    None,
                )

                if comparison:
                    comparison_count = comparison.get("user_count", 0)
                    ws.cell(row, 4, comparison_count)
                    ws.cell(row, 4).number_format = '#,##0'

                    if comparison_count > 0:
                        change = (
                            (metric.get("user_count", 0) - comparison_count)
                            / comparison_count
                        ) * 100
                        ws.cell(row, 5, round(change, 2))
                        ws.cell(row, 5).number_format = '0.00"%"'

            row += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _add_campaigns_sheet(
        self,
        wb: Workbook,
        campaign_metrics: List[Dict],
        comparison_campaigns: Optional[List[Dict]],
    ) -> None:
        """Add campaigns data sheet"""
        ws = wb.create_sheet("Campaigns")

        # Headers
        headers = [
            "Campaign Name",
            "Country",
            "Channel",
            "Type",
            "Sent",
            "Delivered",
            "Opens",
            "Clicks",
            "Unsubscribes",
            "Bounced",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = self._get_header_style()
            cell.fill = self._get_section_header_fill()

        # Data rows
        row = 2
        for metric in campaign_metrics:
            ws.cell(row, 1, metric.get("campaign_name"))
            ws.cell(row, 2, metric.get("country"))
            ws.cell(row, 3, metric.get("channel"))
            ws.cell(row, 4, metric.get("campaign_type"))
            ws.cell(row, 5, metric.get("sent", 0))
            ws.cell(row, 6, metric.get("delivered", 0))
            ws.cell(row, 7, metric.get("open", 0))
            ws.cell(row, 8, metric.get("click", 0))
            ws.cell(row, 9, metric.get("unsubscribe", 0))
            ws.cell(row, 10, metric.get("bounced", 0))

            # Number formatting
            for col in range(5, 11):
                ws.cell(row, col).number_format = '#,##0'

            row += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
